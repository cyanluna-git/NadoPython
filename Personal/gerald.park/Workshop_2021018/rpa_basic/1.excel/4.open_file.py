from openpyxl import load_workbook

wb = load_workbook=("D:\WorkPlace\git_repo\NadoPython\sample.xlsx")
ws = wb.active

for x in range(1,11):
    for y in range(1,11):
        print(ws.cell(row=x,column=y).value, end=" ")
    print()

for x in range(1, ws.max_row +1):
    for y in range(1, ws.max_colume+1):
        print(ws.cell(row=x,colume=y).value, end=" ")
    print


