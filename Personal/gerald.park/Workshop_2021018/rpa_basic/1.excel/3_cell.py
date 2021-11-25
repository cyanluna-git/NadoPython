from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NodoSheet"

#A1 셀에 1이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

# 값읽기 
print(ws["B1"].value)

# row = 1,2,3 ... 
# colume = A(1), B(2), C(3), ... 
print(ws.cell(row=1,column=1).value)
print(ws.cell(row=1,column=2).value)
ws.cell(row=1, column=1)

from random import *
# 반복문 
index = 0
for x in range(1,11):
    for y in range(1,11):
        #ws.cell(row=x,column=y,value=randint(0,100))
        ws.cell(row=x, column = y,value=index)
        index += 1
C = ws.cell(column=3,row=1,value=10)
print(C.value)
wb.save("sample.xlsx")
wb.close()
