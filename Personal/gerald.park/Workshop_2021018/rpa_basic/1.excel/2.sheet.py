from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() #새로운 시트 만들기 

ws.title = "Mysheet"
ws.sheet_properties.tabColor = "ff66ff"


ws1= wb.create_sheet("Your Sheet")
ws2= wb.create_sheet("NewSheet",2) #2번째 인덱스에 시트 생성 

new_ws = wb["NewSheet"]
print(wb.sheetnames) 

#시트 복사 
new_ws["A1"]="Test"
target = wb.copy_worksheet(new_ws)
target.tite = "copied sheet"

wb.save("sample.xlsx")
wb.close()