# 참조 사이트 : https://greeksharifa.github.io/references/2020/10/30/python-selenium-usage/

from googletrans import Translator
import openpyxl

# 엑셀파일 로드
file_name = r'C:\Users\ParkHM\Desktop\Translate.xlsx'
wb = openpyxl.load_workbook(filename=file_name, data_only=True)
ws = wb.active
data = []
data2 = []

row_cnt = ws.max_row

# 엑셀 내용을 data list 에 저장
for col in range(3,5):
    for row in range(1,row_cnt+1):
        data.append(ws.cell(row=row, column=col).value)


# 구글 번역기
i=0
trans = Translator()
for text in data:
    i+=1
    print("Progress..",i,"/",len(data)) # 진행현황 확인을 위해 표시
    result = trans.translate(str(text), dest='ko')
    data2.append(result.text)
    
# data2 list 를 엑셀에 다시 저장
for col in range (6,8):
    for row in range(1,row_cnt+1):
        ws.cell(row=row, column=col).value = data2[(col-6)*(row_cnt)+row-1]

wb.save(r'C:\Users\ParkHM\Desktop\Translate.xlsx')



 