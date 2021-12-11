from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active


# 한 줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11): # 총 10개 데이터 넣기
    ws.append([i, randint(0,100), randint(0,100)])


col_B = ws["B"] # B column 만 가지고 오기
# print(col_B)
# for cell in col_B:
#     print(cell.value)


# col_range = ws["B:C"] # B 와 C column 함께 가지고 오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# row_title = ws[1] # 첫번째 row만 가지고 오기
# for cell in row_title:
#     print(cell.value)

row_range = ws[2:6] # 2~6번째 줄 가지고 오기, 6 포함임에 주의
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

wb.save("sample.xlsx")